"""scripts/_phi_universal_xtrad_sizing.py
==========================================
Discovery / sizing script for the cross-tradition universal Phi
experiment (exp109; the genuine Quran-distinctiveness test that pairs
with the F55 cross-tradition results F59/F60/F61/F62).

Question (HONEST): in a STRICTLY UNIVERSAL 5-D structural feature space
that contains NO Arabic morphology, NO language-specific connective
lists, and NO Quran-specific metadata, where does the Quran rank
against 11 other corpora (6 Arabic peers + Hebrew Tanakh + Greek NT
+ Pali DN/MN + Avestan Yasna)?

Universal 5-D feature set (locked at PREREG seal):
  F1 : VL_CV               (verse-length coefficient of variation; words)
  F2 : p_max(end_letter)   (top-frequency end-letter / total verses)
  F3 : H_EL                (Shannon entropy of end-letter distribution; bits)
  F4 : bigram_distinct_ratio (n_distinct_bigrams / n_bigram_positions on skeleton)
  F5 : gzip_efficiency     (gzip(skeleton) / raw(skeleton); lower = more redundant)

Per corpus, take MEDIAN over units (chapter / sutta / yasna / surah / poem).
Then ask: does the Quran sit at any extremum (rank 1 or rank 12) on each
feature, and across-features, how does it rank by joint Mahalanobis distance
to the centroid of all 12 corpora?

This is a NULL-RESULT-PUBLISHABLE experiment. We commit to reporting
honestly whether the Quran ranks #1 (real cross-tradition Quran-finding)
or somewhere middle (null result that bounds the universality claim).

NO COMPRESSION CALLS to PREREG-locked hashes; uses gzip purely for the
F5 feature itself, which is part of the locked feature definition.
"""
from __future__ import annotations

import gzip
import json
import math
import sys
import time
import unicodedata
import html
import re
from collections import Counter
from pathlib import Path

import numpy as np

_HERE = Path(__file__).resolve().parent
_ROOT = _HERE.parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))


# ============================================================================
# Per-tradition NORMALISERS (locked, copied from exp105-108 + features.py)
# ============================================================================

# --- Arabic 28-letter skeleton (Quran + Arabic peers) -----------------------
_AR_DIAC = re.compile(r"[\u0610-\u061A\u064B-\u065F\u06D6-\u06ED\u0670]")
_AR_TATWEEL = "\u0640"
_ARABIC_LETTERS_28 = "ابتثجحخدذرزسشصضطظعغفقكلمنهوي"
_ARABIC_SET = set(_ARABIC_LETTERS_28)


def _normalise_arabic(s: str) -> str:
    s = s.replace(_AR_TATWEEL, "")
    s = _AR_DIAC.sub("", s)
    # Fold alif variants and ya/ta-marbuta to canonical 28
    s = (s.replace("\u0622", "\u0627")  # ا
          .replace("\u0623", "\u0627")  # ا
          .replace("\u0625", "\u0627")  # ا
          .replace("\u0671", "\u0627")  # ا
          .replace("\u0649", "\u064a")  # ي
          .replace("\u0629", "\u0647")) # ه (ta-marbuta to ha)
    return "".join(c for c in s if c in _ARABIC_SET)


# --- Hebrew 22-consonant skeleton -------------------------------------------
_HEBREW_CONS_22 = "אבגדהוזחטיכלמנסעפצקרשת"
_HEBREW_SET = set(_HEBREW_CONS_22)
_HEBREW_FINAL_FOLD = {"ך": "כ", "ם": "מ", "ן": "נ", "ף": "פ", "ץ": "צ"}
_HEBREW_DIAC = re.compile(
    r"[\u0591-\u05BD\u05BF\u05C1-\u05C2\u05C4-\u05C5\u05C7]"
)


def _normalise_hebrew(s: str) -> str:
    s = _HEBREW_DIAC.sub("", s)
    for k, v in _HEBREW_FINAL_FOLD.items():
        s = s.replace(k, v)
    return "".join(c for c in s if c in _HEBREW_SET)


# --- Greek 24-letter skeleton (sigma-fold + diacritic strip) ---------------
_GREEK_LETTERS_24 = "αβγδεζηθικλμνξοπρστυφχψω"
_GREEK_SET = set(_GREEK_LETTERS_24)
_GREEK_SIGMA_FOLD = {"ς": "σ", "ϲ": "σ", "Ϲ": "σ"}


def _normalise_greek(s: str) -> str:
    for k, v in _GREEK_SIGMA_FOLD.items():
        s = s.replace(k, v)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.casefold()
    return "".join(c for c in s if c in _GREEK_SET)


# --- Pali 31-letter IAST skeleton ------------------------------------------
_PALI_ALPHABET_31 = (
    "a\u0101i\u012Bu\u016Beo"
    "kg\u1E45c"
    "j\u00F1"
    "\u1E6D\u1E0D\u1E47"
    "tdn"
    "pbm"
    "yrl\u1E37"
    "vsh"
    "\u1E41"
)
_PALI_SET = set(_PALI_ALPHABET_31)


def _normalise_pali(s: str) -> str:
    s = unicodedata.normalize("NFC", s)
    s = s.casefold()
    s = s.replace("\u1E43", "\u1E41")
    return "".join(c for c in s if c in _PALI_SET)


# --- Avestan 26-letter Latin transliteration -------------------------------
_AV_ALPHABET_26 = "abcdefghijklmnopqrstuvwxyz"
_AV_SET = set(_AV_ALPHABET_26)
_TAG_RE = re.compile(r"<[^>]+>")


def _normalise_avestan(s: str) -> str:
    if not s:
        return ""
    s = html.unescape(s)
    s = _TAG_RE.sub(" ", s)
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.casefold()
    return "".join(c for c in s if c in _AV_SET)


# ============================================================================
# UNIT loaders per tradition (returns list of (label, verses_list, n_words))
# ============================================================================

# --- Quran ------------------------------------------------------------------
def _load_quran() -> list[dict]:
    p = _ROOT / "data" / "corpora" / "ar" / "quran_bare.txt"
    surahs: dict[int, list[str]] = {}
    for line in p.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        parts = line.split("|", 2)
        if len(parts) != 3:
            continue
        sid = int(parts[0])
        surahs.setdefault(sid, []).append(parts[2].strip())
    out: list[dict] = []
    for sid in sorted(surahs):
        verses = surahs[sid]
        out.append({
            "corpus": "quran",
            "label": f"Q:{sid:03d}",
            "verses": verses,
            "normaliser": "arabic",
        })
    return out


# --- Arabic peer corpora ---------------------------------------------------
def _load_arabic_peers() -> dict[str, list[dict]]:
    """Returns {corpus_name: list_of_units}.
    Use raw_loader for poetry / quran / arabic_bible / ksucca / hindawi."""
    out: dict[str, list[dict]] = {}
    try:
        import csv
        # Poetry CSV split into 7-word chunks per the locked convention
        poetry_csv = _ROOT / "data" / "corpora" / "ar" / "poetry_raw.csv"
        ERA_MAP = {
            "العصر الجاهلي": "poetry_jahili",
            "المخضرمون": "poetry_jahili",
            "العصر الاسلامي": "poetry_islami",
            "العصر الاموي": "poetry_islami",
            "العصر العباسي": "poetry_abbasi",
        }
        for c in ("poetry_jahili", "poetry_islami", "poetry_abbasi"):
            out[c] = []
        with poetry_csv.open("r", encoding="utf-8", errors="replace") as f:
            rd = csv.DictReader(f)
            for i, row in enumerate(rd):
                era_ar = (row.get("era") or "").strip()
                corpus = ERA_MAP.get(era_ar)
                if not corpus:
                    continue
                poem = (row.get("poem") or "").strip()
                if not poem:
                    continue
                ws = poem.split()
                if len(ws) < 35:  # require min 5 hemistichs
                    continue
                verses = [" ".join(ws[k:k+7]) for k in range(0, len(ws), 7)]
                out[corpus].append({
                    "corpus": corpus,
                    "label": f"{corpus}:{row.get('poem_id') or i}",
                    "verses": verses,
                    "normaliser": "arabic",
                })
    except Exception as e:
        print(f"# WARN: poetry load failed: {e}", file=sys.stderr)

    # Hindawi (modern Arabic prose) — split by lines, group into "units" of
    # ~50 lines each (rough chapter analog).
    try:
        hindawi = _ROOT / "data" / "corpora" / "ar" / "hindawi.txt"
        lines = [l.strip() for l in hindawi.read_text(encoding="utf-8").splitlines() if l.strip()]
        out["hindawi"] = []
        chunk = 50
        for i in range(0, len(lines), chunk):
            verses = lines[i:i+chunk]
            if len(verses) < 10:
                continue
            out["hindawi"].append({
                "corpus": "hindawi",
                "label": f"hindawi:{i//chunk:04d}",
                "verses": verses,
                "normaliser": "arabic",
            })
    except Exception as e:
        print(f"# WARN: hindawi load failed: {e}", file=sys.stderr)

    # Ksucca (classical Arabic ksucca-style prose)
    try:
        ksucca = _ROOT / "data" / "corpora" / "ar" / "ksucca.txt"
        text = ksucca.read_text(encoding="utf-8")
        # Split on dhikr markers as chapter analog
        DHIKR = re.compile(r"^\s*ذكر\b", re.MULTILINE)
        positions = [m.start() for m in DHIKR.finditer(text)] + [len(text)]
        out["ksucca"] = []
        for i in range(len(positions) - 1):
            chunk = text[positions[i]:positions[i+1]].strip()
            if not chunk:
                continue
            verses = [v.strip() for v in chunk.split("\n") if v.strip()]
            if len(verses) < 5:
                continue
            out["ksucca"].append({
                "corpus": "ksucca",
                "label": f"ksucca:{i:04d}",
                "verses": verses,
                "normaliser": "arabic",
            })
    except Exception as e:
        print(f"# WARN: ksucca load failed: {e}", file=sys.stderr)

    # Arabic Bible (Smith Van Dyke; XLSX). Row 5 has headers
    # 'Verse ID, Book Name, Book Number, Chapter, Verse, Text'.
    # Aggregate to (Book Number, Chapter) granularity.
    try:
        import openpyxl
        wb = openpyxl.load_workbook(
            _ROOT / "data" / "corpora" / "ar" / "arabic_bible.xlsx",
            read_only=True, data_only=True,
        )
        ws = wb[wb.sheetnames[0]]
        chapters_ab: dict[tuple[int, int], list[str]] = {}
        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if header_row is None:
                if row and "Book Name" in (str(c) for c in row if c):
                    header_row = i
                continue
            try:
                book_num = int(row[2])
                chap = int(row[3])
                txt = (row[5] or "").strip()
            except (ValueError, TypeError, IndexError):
                continue
            if not txt:
                continue
            chapters_ab.setdefault((book_num, chap), []).append(txt)
        wb.close()
        out["arabic_bible"] = []
        for (bk, ch), verses in chapters_ab.items():
            if len(verses) < 5:
                continue
            out["arabic_bible"].append({
                "corpus": "arabic_bible",
                "label": f"AB:{bk:02d}:{ch:03d}",
                "verses": verses,
                "normaliser": "arabic",
            })
    except Exception as e:
        print(f"# WARN: arabic_bible load failed: {e}", file=sys.stderr)
    return out


# --- Hebrew Tanakh (WLC) ----------------------------------------------------
def _load_hebrew_tanakh() -> list[dict]:
    """Use the locked Hebrew Tanakh loader from exp104c (validated for
    F53 + F55 cross-tradition runs)."""
    try:
        from experiments.exp104_F53_tanakh_pilot.run import _load_tanakh_chapters
    except Exception as e:
        print(f"# WARN: Hebrew loader unavailable: {e}", file=sys.stderr)
        return []
    chapters = _load_tanakh_chapters()
    out: list[dict] = []
    for ch in chapters:
        verses = ch.get("verses_raw") or ch.get("verses") or []
        if len(verses) < 5:
            continue
        out.append({
            "corpus": "hebrew_tanakh",
            "label": f"{ch.get('book', '?')}:{ch.get('chapter', '?')}",
            "verses": verses,
            "normaliser": "hebrew",
        })
    return out


# --- Greek NT (OpenGNT v3.3) -----------------------------------------------
def _load_greek_nt() -> list[dict]:
    try:
        from experiments.exp104d_F53_mark1.run import _load_greek_nt_chapters
    except Exception as e:
        print(f"# WARN: Greek loader unavailable: {e}", file=sys.stderr)
        return []
    chapters = _load_greek_nt_chapters()
    out: list[dict] = []
    for ch in chapters:
        # Greek loader uses 'verses_raw' (not 'verses')
        verses = ch.get("verses_raw") or ch.get("verses") or []
        if not verses or len(verses) < 5:
            continue
        out.append({
            "corpus": "greek_nt",
            "label": f"book{ch['book']}:ch{ch['chapter']}",
            "verses": verses,
            "normaliser": "greek",
        })
    return out


# --- Pali DN+MN -------------------------------------------------------------
def _load_pali() -> list[dict]:
    try:
        from experiments.exp107_F55_dn1_bigram.run import _load_pali_suttas
    except Exception as e:
        print(f"# WARN: Pali loader unavailable: {e}", file=sys.stderr)
        return []
    suttas = _load_pali_suttas()
    out: list[dict] = []
    for s in suttas:
        # Split text_raw on segment-key boundaries: SuttaCentral keys are
        # "dnX:S.P.SS"; we approximate verses by splitting on sentence
        # punctuation in the raw text since segments are inside strings.
        # Simpler: each sutta JSON has values; reload to get verses-as-values.
        f = Path(s["file"])
        try:
            data = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            continue
        verses = [str(v).strip() for v in data.values() if str(v).strip()]
        if len(verses) < 5:
            continue
        out.append({
            "corpus": "pali",
            "label": s["sutta_id"],
            "verses": verses,
            "normaliser": "pali",
        })
    return out


# --- Avestan Yasna (Geldner / Avesta.org) -----------------------------------
def _load_avestan() -> list[dict]:
    try:
        from experiments.exp108_F55_y28_bigram.run import _load_avestan_yasnas
    except Exception as e:
        print(f"# WARN: Avestan loader unavailable: {e}", file=sys.stderr)
        return []
    yasnas = _load_avestan_yasnas()
    out: list[dict] = []
    for y in yasnas:
        # text_raw is " "-joined verses; we split back via the loader's
        # verse capture: re-parse the source file's <DD> blocks.
        # Simpler: split text_raw at multiple-space gaps as verse-boundaries.
        # The loader joined verses with " " so this is lossy. Acceptable for
        # the sizing diagnostic; a finer parse can be done in the run.py.
        verses = [v.strip() for v in re.split(r"\s{2,}|(?<=\.)\s+(?=[A-Za-z])", y["text_raw"]) if v.strip()]
        if len(verses) < 5:
            verses = [v.strip() for v in y["text_raw"].split(". ") if v.strip()]
        if len(verses) < 5:
            continue
        out.append({
            "corpus": "avestan_yasna",
            "label": f"y{y['yasna']:02d}",
            "verses": verses,
            "normaliser": "avestan",
        })
    return out


# ============================================================================
# Universal 5-D feature extractor (no Arabic morphology, no per-language stops)
# ============================================================================

NORMALISERS = {
    "arabic":   _normalise_arabic,
    "hebrew":   _normalise_hebrew,
    "greek":    _normalise_greek,
    "pali":     _normalise_pali,
    "avestan":  _normalise_avestan,
}


def _terminal_letter(verse: str, normaliser_name: str) -> str:
    """Return the LAST alphabet-letter of the normalised verse, or ''."""
    norm = NORMALISERS[normaliser_name](verse)
    return norm[-1] if norm else ""


def _features_universal(unit: dict) -> dict:
    """Compute the 5 universal features for one unit."""
    norm_fn = NORMALISERS[unit["normaliser"]]
    verses = unit["verses"]

    # --- F1: VL_CV (verse-length CV, in WORDS) ------------------------------
    word_lens = [len(v.split()) for v in verses if v]
    if len(word_lens) < 2 or sum(word_lens) == 0:
        vl_cv = 0.0
    else:
        arr = np.array(word_lens, dtype=float)
        vl_cv = float(arr.std(ddof=1) / arr.mean()) if arr.mean() > 0 else 0.0

    # --- F2 / F3: end-letter distribution -----------------------------------
    finals = [_terminal_letter(v, unit["normaliser"]) for v in verses]
    finals = [f for f in finals if f]  # drop empties
    if finals:
        c = Counter(finals)
        total = sum(c.values())
        p_max = max(c.values()) / total
        h_el = 0.0
        for n in c.values():
            p = n / total
            if p > 0:
                h_el -= p * math.log2(p)
    else:
        p_max = 0.0
        h_el = 0.0

    # --- Skeleton-based features --------------------------------------------
    skeleton = "".join(norm_fn(v) for v in verses)
    n = len(skeleton)

    # --- F4: bigram_distinct_ratio ------------------------------------------
    if n < 2:
        bigram_distinct_ratio = 0.0
    else:
        bigrams = [skeleton[i:i+2] for i in range(n-1)]
        bigram_distinct_ratio = len(set(bigrams)) / max(1, len(bigrams))

    # --- F5: gzip_efficiency ------------------------------------------------
    if n < 50:
        gzip_efficiency = 1.0
    else:
        raw = skeleton.encode("utf-8")
        compressed = gzip.compress(raw, compresslevel=9)
        gzip_efficiency = len(compressed) / max(1, len(raw))

    return {
        "label": unit["label"],
        "n_verses": len(verses),
        "n_words": sum(word_lens),
        "n_letters_skeleton": n,
        "VL_CV": vl_cv,
        "p_max": p_max,
        "H_EL": h_el,
        "bigram_distinct_ratio": bigram_distinct_ratio,
        "gzip_efficiency": gzip_efficiency,
    }


# ============================================================================
# Main: load all 12 corpora, extract features, rank
# ============================================================================

def main() -> int:
    t0 = time.time()
    print("# _phi_universal_xtrad_sizing.py  --  exp109 sizing")
    print()

    corpora: dict[str, list[dict]] = {}

    print("# Loading corpora ...")
    corpora["quran"] = _load_quran()
    print(f"#   quran: {len(corpora['quran'])} surahs")

    arabic_peers = _load_arabic_peers()
    for name, units in arabic_peers.items():
        corpora[name] = units
        print(f"#   {name}: {len(units)} units")

    corpora["hebrew_tanakh"] = _load_hebrew_tanakh()
    print(f"#   hebrew_tanakh: {len(corpora['hebrew_tanakh'])} chapters")

    corpora["greek_nt"] = _load_greek_nt()
    print(f"#   greek_nt: {len(corpora['greek_nt'])} chapters")

    corpora["pali"] = _load_pali()
    print(f"#   pali: {len(corpora['pali'])} suttas")

    corpora["avestan_yasna"] = _load_avestan()
    print(f"#   avestan_yasna: {len(corpora['avestan_yasna'])} yasnas")

    # Filter: drop corpora with < 10 units (insufficient for stable median)
    corpora = {k: v for k, v in corpora.items() if len(v) >= 10}
    print()
    print(f"# Surviving corpora ({len(corpora)}): {list(corpora.keys())}")
    print()

    # --- Compute features per unit ------------------------------------------
    feats: dict[str, list[dict]] = {}
    for name, units in corpora.items():
        feats[name] = []
        for u in units:
            try:
                f = _features_universal(u)
                feats[name].append(f)
            except Exception as e:
                print(f"# WARN: {name}/{u.get('label')} failed: {e}",
                      file=sys.stderr)
        print(f"# Features for {name}: {len(feats[name])} units")

    # --- Per-corpus median feature vector -----------------------------------
    print()
    print("# === Per-corpus median feature vector ===")
    medians: dict[str, dict] = {}
    feature_names = ["VL_CV", "p_max", "H_EL",
                     "bigram_distinct_ratio", "gzip_efficiency"]
    for name, fs in feats.items():
        m: dict[str, float] = {}
        for k in feature_names:
            vals = [f[k] for f in fs]
            m[k] = float(np.median(vals)) if vals else float("nan")
        m["n_units"] = len(fs)
        medians[name] = m

    # Print as table
    cols = ["corpus      ", "n_units"] + feature_names
    print("# " + "  ".join(f"{c:>22s}" for c in cols))
    for name, m in medians.items():
        cells = [f"{name:<12s}", f"{m['n_units']:>5d}"]
        for k in feature_names:
            cells.append(f"{m[k]:>22.4f}")
        print("# " + "  ".join(f"{c:>22s}" for c in cells))

    # --- Ranking per feature ------------------------------------------------
    print()
    print("# === Per-feature corpus ranking (1 = lowest, N = highest) ===")
    rankings: dict[str, dict[str, int]] = {}
    n_corp = len(medians)
    for k in feature_names:
        sorted_corp = sorted(medians.keys(), key=lambda c: medians[c][k])
        for rank, c in enumerate(sorted_corp, start=1):
            rankings.setdefault(c, {})[k] = rank
    cols = ["corpus      "] + feature_names
    print("# " + "  ".join(f"{c:>22s}" for c in cols))
    for name in medians.keys():
        cells = [f"{name:<12s}"]
        for k in feature_names:
            r = rankings[name][k]
            cells.append(f"{r:>22d}")
        print("# " + "  ".join(f"{c:>22s}" for c in cells))

    # --- Quran's rank profile -----------------------------------------------
    print()
    print("# === Quran's rank profile across the 5 universal features ===")
    if "quran" in rankings:
        n_extreme = sum(1 for k in feature_names
                        if rankings["quran"][k] in (1, n_corp))
        ranks_str = ", ".join(f"{k}=#{rankings['quran'][k]}/{n_corp}"
                              for k in feature_names)
        print(f"#   Quran ranks: {ranks_str}")
        print(f"#   Quran is at extremum (rank 1 or {n_corp}) on {n_extreme}/5 features")
    else:
        print("#   Quran not in surviving corpora!")

    # --- Save sizing receipt ------------------------------------------------
    out_dir = _ROOT / "results" / "auxiliary"
    out_dir.mkdir(parents=True, exist_ok=True)
    receipt = {
        "schema": "phi_universal_xtrad_sizing_v1",
        "filed_as": "exploratory_diagnostic",
        "purpose": ("Sanity check before exp109_phi_universal_cross_tradition "
                    "PREREG. Cross-tradition Quran-distinctiveness on a "
                    "strictly universal 5-D structural feature space."),
        "features": feature_names,
        "n_corpora": n_corp,
        "n_units_per_corpus": {c: medians[c]["n_units"] for c in medians},
        "medians": medians,
        "ranks": rankings,
        "wall_time_s": time.time() - t0,
    }
    out_path = out_dir / "_phi_universal_xtrad_sizing.json"
    out_path.write_text(json.dumps(receipt, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    print()
    print(f"# wall-time: {time.time() - t0:.1f}s")
    print(f"# receipt: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
